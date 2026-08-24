"""
Tier 3 dataset: a small .xlsx, generated rather than committed — and the reason is NOT size.

    python datasets/generated/make_messy_excel.py

WHY GENERATED WHEN IT IS ONLY FOURTEEN ROWS
Every other tier-3 file is here because a million rows do not belong in git. This one is tiny
and is still generated, because of tier 2's actual rationale: a messy dataset is committed as
TEXT so a reviewer can see the defect in a diff. An .xlsx is a zip of XML. Committing one
would keep the letter of tier 2 and lose the entire point of it — nobody can review a binary,
and six months from now nobody would remember which cells were deliberately broken.

So the defects are AUTHORED HERE, in reviewable Python, and the binary is built from them. The
four constants below are the episode's syllabus, in a diff, permanently.

WHY THIS FILE EXISTS AT ALL
Series 2 · 02 (Excel Files) needs defects that CANNOT be expressed in a CSV, which is exactly
why that episode is not a footnote on Series 2 · 00. A learner who has done the CSV episodes
has met none of the four below.

WHY A FIXED SEED
Same rule as make_big_orders.py: koans assert on exact numbers, so every learner must get a
byte-identical file. `random.Random(SEED)`, never the global module.

NEEDS openpyxl, which is in tests/python/requirements.txt because the Excel koans need it too.
"""

import os
import random
from datetime import date

SEED = 20260824          # same seed discipline as make_big_orders.py — do not change casually
OUT = os.path.dirname(os.path.abspath(__file__))

# ── THE FOUR DEFECTS, which are the episode's syllabus ─────────────────────
#
#  1. TITLE ROW ABOVE THE HEADER. The real header is row 3; rows 1-2 are a title and a blank.
#     Default read_excel gives you columns called `Unnamed: 1`. Fix: skiprows / header=2.
TITLE = "Q3 Orders — Confidential"
#
#  2. MERGED CELLS. "Customer" spans two columns (name + code). openpyxl reports the value in
#     the top-left cell of the span and None everywhere else, so one column arrives unnamed
#     and the other arrives all-NaN.
MERGE_SPAN = "B3:C3"
#
#  3. A DATE STORED AS ITS SERIAL NUMBER. Excel counts days from 1899-12-30. Written as a
#     plain integer with no date format, so it reads back as 45123 — a perfectly plausible
#     order ID, which is what makes it dangerous rather than merely annoying.
EXCEL_EPOCH = date(1899, 12, 30)
SERIAL_ROWS = [3, 7, 11]          # 0-based into the data rows
#
#  4. A NUMBER STORED AS TEXT. In Excel the leading apostrophe is invisible; in pandas the
#     column arrives as object dtype, so .sum() CONCATENATES instead of adding. This is the
#     worst defect in the file because it produces an answer rather than an error.
TEXT_NUMBER_ROWS = [1, 5, 9, 12]  # 0-based into the data rows

CUSTOMERS = [
    ("Alfreds Futterkiste", "ALFKI"), ("Ana Trujillo", "ANATR"),
    ("Antonio Moreno", "ANTON"), ("Around the Horn", "AROUT"),
    ("Blauer See Delikatessen", "BLAUS"), ("Blondel père et fils", "BLONP"),
    ("Bólido Comidas", "BOLID"),
]


def main():
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit(
            "openpyxl is not installed here.\n"
            "  pip install openpyxl   — or run this inside the koan image, which has it.")

    rnd = random.Random(SEED)
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Defect 1 — a title where a header should be.
    ws["A1"] = TITLE
    # row 2 stays empty, because that is what real exports look like

    header = ["OrderID", "Customer", "Code", "OrderDate", "Freight", "Country"]
    for i, h in enumerate(header, start=1):
        ws.cell(row=3, column=i, value=h)

    # Defect 2 — merge the two Customer columns so the second header vanishes.
    ws.merge_cells(MERGE_SPAN)

    for r in range(14):
        name, code = CUSTOMERS[r % len(CUSTOMERS)]
        d = date(2026, 7, 1 + rnd.randrange(0, 28))
        freight = round(rnd.uniform(10.0, 400.0), 2)
        row = 4 + r

        ws.cell(row=row, column=1, value=10248 + r)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=code)

        # Defect 3 — some dates as real dates, some as bare serial numbers.
        if r in SERIAL_ROWS:
            ws.cell(row=row, column=4, value=(d - EXCEL_EPOCH).days)
        else:
            ws.cell(row=row, column=4, value=d).number_format = "yyyy-mm-dd"

        # Defect 4 — some freights as numbers, some as text.
        c = ws.cell(row=row, column=5)
        if r in TEXT_NUMBER_ROWS:
            c.value = f"{freight:.2f}"
            c.data_type = "s"
        else:
            c.value = freight

        ws.cell(row=row, column=6, value=rnd.choice(
            ["Germany", "Mexico", "UK", "France", "Brazil"]))

    for i in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    path = os.path.join(OUT, "orders-export.xlsx")
    wb.save(path)

    # Print the answers a koan author needs, so nobody has to open Excel to write the koans.
    true_freight = sum(
        ws.cell(row=4 + r, column=5).value if r not in TEXT_NUMBER_ROWS
        else float(ws.cell(row=4 + r, column=5).value)
        for r in range(14))
    print(f"wrote {path}  (14 data rows, header on row 3)")
    print(f"  defect 1  title row      : A1 = {TITLE!r}")
    print(f"  defect 2  merged cells   : {MERGE_SPAN}")
    print(f"  defect 3  date serials   : data rows {SERIAL_ROWS} (0-based)")
    print(f"  defect 4  numbers as text: data rows {TEXT_NUMBER_ROWS} (0-based)")
    print(f"  true total freight       : {true_freight:.2f}   <- the koan's right answer")


if __name__ == "__main__":
    main()
